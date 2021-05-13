#-*- coding: utf-8 -*-
# @Author  : wuqi
# @Time    : 2021/5/12 11:05
# @Software: PyCharm
# @File    : xlutils.py
from openpyxl import load_workbook


class xl:
    def __init__(self, path, sheetName):
        self.path = path
        self.sheet = sheetName

    def get_value(self, row, column):
        sheet = load_workbook(self.path)[self.sheet]
        return sheet.cell(row, column).value


if __name__ == '__main__':
    value = xl(r'C:\Users\wuqi14\Desktop\车架号重复车辆.xlsx', 'Sheet1').get_value(1, 1)
    print(value)