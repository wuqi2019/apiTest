#-*- coding: utf-8 -*-
# @Author  : wuqi
# @Time    : 2021/5/10 11:42
# @Software: PyCharm
# @File    : ddt01.py
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
# ws = wb.active
wb = load_workbook(r'C:\Users\wuqi14\Desktop\车架号重复车辆.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for sheet in wb:
    print(sheet.title)

a = s