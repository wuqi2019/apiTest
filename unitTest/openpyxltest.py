#-*- coding: utf-8 -*-
# @Author  : wuqi
# @Time    : 2021/5/11 16:27
# @Software: PyCharm
# @File    : openpyxltest.py
from openpyxl import load_workbook

# 1、打开工作簿
wb = load_workbook(r'C:\Users\wuqi14\Desktop\车架号重复车辆.xlsx')

# 2、指定表单
sheet = wb['Sheet1']

# 3、根据行列值定位单元格
value = sheet.cell(1, 1).value

print(value)

print(sheet.max_row)
print(sheet.max_column)